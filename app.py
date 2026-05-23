"""
Gestor de Finanzas — Backend Flask
Diseño: periódico antiguo, papel envejecido, tipografía clásica serif
"""

from flask import Flask, request, jsonify, render_template_string
from datetime import datetime
import openpyxl, os, re

app = Flask(__name__)

EXCEL_PATH     = os.environ.get("EXCEL_PATH", "Finanzas_Miguel_2026.xlsx")
CLAUDE_API_KEY = os.environ.get("CLAUDE_API_KEY", "")

MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
            7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

REGLAS = {
    "Comida":      ["mercadona","lidl","aldi","carrefour","dia","eroski","supermercado",
                    "fruteria","panaderia","pan","comida","cena","almuerzo","desayuno",
                    "cafe","bar","restaurante","pizza","burger","mcdonalds","kfc","kebab",
                    "delivery","glovo","uber eats","just eat","sushi","bocadillo",
                    "fruta","verdura","carne","pescado"],
    "Transporte":  ["gasolina","combustible","repsol","bp","cepsa","parking","aparcar",
                    "autobus","metro","cercanias","renfe","tren","taxi","uber","cabify",
                    "blablacar","peaje","itv","seguro coche","coche","moto","bici","patinete"],
    "Hogar":       ["alquiler","hipoteca","piso","casa","habitacion","luz","electricidad",
                    "agua","gas natural","factura gas","wifi","internet","fibra","movistar",
                    "vodafone","orange","comunidad","seguro hogar","ikea","leroy",
                    "bricolaje","mueble","reparacion"],
    "Salud":       ["farmacia","medicamento","medicina","doctor","medico","dentista",
                    "hospital","clinica","consulta","gym","gimnasio","proteina","suplemento",
                    "seguro medico","fisio","fisioterapia"],
    "Ocio":        ["netflix","spotify","amazon prime","disney","hbo","youtube premium",
                    "suscripcion","cine","teatro","concierto","viaje","vuelo","hotel",
                    "airbnb","booking","vacaciones","videojuego","steam","libro","amazon",
                    "fnac","copa","cerveza","discoteca","ocio"],
    "Inversiones": ["bolsa","accion","etf","fondo","inversion","crypto","bitcoin",
                    "ethereum","trading","degiro","ahorro","deposito","plan de pensiones",
                    "indexa","myinvestor","finizens"],
    "Trabajo":     ["material","herramienta","ordenador","portatil","software","licencia",
                    "curso","formacion","libro tecnico","office","adobe","asesoria",
                    "gestor","notario","autonomo","cuota autonomos"],
    "Ingresos":    ["salario","nomina","sueldo","paga","ingreso","transferencia recibida",
                    "freelance","cliente","factura cobrada","devolucion","reembolso",
                    "premio","bonus","extra"],
}

def clasificar_por_reglas(descripcion):
    desc = descripcion.lower()
    for categoria, palabras in REGLAS.items():
        for palabra in palabras:
            if re.search(r'\b' + re.escape(palabra) + r'\b', desc):
                return categoria, ("Ingreso" if categoria == "Ingresos" else "Gasto")
    for categoria, palabras in REGLAS.items():
        for palabra in palabras:
            if palabra in desc:
                return categoria, ("Ingreso" if categoria == "Ingresos" else "Gasto")
    return "Otros", "Gasto"

def clasificar_con_claude(descripcion):
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
        cats = ", ".join(list(REGLAS.keys()) + ["Otros"])
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=50,
            messages=[{"role":"user","content":
                f'Clasifica esta transacción: "{descripcion}"\n'
                f'Categorías: {cats}\n'
                f'Responde SOLO: CATEGORIA|Ingreso o CATEGORIA|Gasto'}])
        resp = msg.content[0].text.strip()
        if "|" in resp:
            cat, tipo = resp.split("|", 1)
            cat, tipo = cat.strip(), tipo.strip()
            if cat in list(REGLAS.keys()) + ["Otros"] and tipo in ["Ingreso","Gasto"]:
                return cat, tipo
    except Exception:
        pass
    return clasificar_por_reglas(descripcion)

def clasificar(descripcion):
    if CLAUDE_API_KEY:
        return clasificar_con_claude(descripcion)
    return clasificar_por_reglas(descripcion)

def añadir_al_excel(descripcion, monto, categoria, tipo, notas=""):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        return -1
    ws = wb["Transacciones"]
    next_row = 5
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:
                next_row = cell.row + 1
    hoy = datetime.now()
    ws[f"B{next_row}"] = descripcion
    ws[f"C{next_row}"] = monto
    ws[f"D{next_row}"] = hoy.strftime("%d/%m/%Y")
    ws[f"E{next_row}"] = categoria
    ws[f"F{next_row}"] = tipo
    ws[f"G{next_row}"] = MESES_ES[hoy.month]
    ws[f"H{next_row}"] = notas
    wb.save(EXCEL_PATH)
    return next_row


HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>MN · Ledger</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,400;0,700;0,900;1,400;1,700&family=UnifrakturMaguntia&family=IM+Fell+English:ital@0;1&display=swap');

:root {
  --paper:   #F0E6C8;
  --paper2:  #E8D9B0;
  --paper3:  #DDD0A0;
  --ink:     #0F0A04;
  --ink2:    #2A1F0E;
  --ink3:    #4A3820;
  --ink4:    #6B5535;
  --faded:   #8C7A5A;
  --red:     #7A1515;
  --green:   #1A4A2A;
  --rule:    #2A1F0E;
}

* { box-sizing: border-box; margin: 0; padding: 0; }

body {
  font-family: 'IM Fell English', Georgia, serif;
  background: var(--paper);
  background-image:
    radial-gradient(ellipse at 20% 10%, rgba(180,150,80,0.15) 0%, transparent 50%),
    radial-gradient(ellipse at 80% 90%, rgba(120,90,40,0.12) 0%, transparent 50%);
  min-height: 100vh;
  color: var(--ink);
  padding-bottom: 60px;
}

/* ── MASTHEAD ── */
.masthead {
  border-bottom: 3px double var(--rule);
  padding: 12px 16px 10px;
  background: var(--paper);
  position: sticky;
  top: 0;
  z-index: 50;
}

.mast-top {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 6px;
}

.mast-logo {
  display: flex;
  align-items: center;
  gap: 8px;
}

.logo-svg { width: 38px; height: 38px; }

.mast-brand {
  font-family: 'Playfair Display', serif;
  font-size: 22px;
  font-weight: 900;
  color: var(--ink);
  letter-spacing: 2px;
  line-height: 1;
}

.mast-sub {
  font-family: 'IM Fell English', serif;
  font-style: italic;
  font-size: 10px;
  color: var(--ink4);
  letter-spacing: 1px;
}

.mast-date {
  font-family: 'IM Fell English', serif;
  font-size: 10px;
  color: var(--ink3);
  text-align: right;
  font-style: italic;
}

.mast-rule {
  border: none;
  border-top: 1px solid var(--rule);
  margin: 6px 0 0;
}

.mast-headline {
  text-align: center;
  font-family: 'Playfair Display', serif;
  font-size: 11px;
  font-weight: 400;
  letter-spacing: 3px;
  text-transform: uppercase;
  color: var(--ink3);
  margin-top: 4px;
}

/* ── STATS / LEDGER HEADER ── */
.ledger-bar {
  display: grid;
  grid-template-columns: 1fr 1px 1fr 1px 1fr;
  margin: 12px 16px;
  border: 1.5px solid var(--rule);
  background: var(--paper2);
}

.ledger-col {
  padding: 12px 8px;
  text-align: center;
  transition: background 0.2s ease;
  cursor: default;
}

.ledger-col:hover { background: var(--paper3); }

.ledger-divider {
  background: var(--rule);
  width: 1px;
}

.ledger-val {
  font-family: 'Playfair Display', serif;
  font-size: 20px;
  font-weight: 700;
  color: var(--ink);
  line-height: 1;
}

.ledger-lbl {
  font-family: 'IM Fell English', serif;
  font-size: 9px;
  font-style: italic;
  color: var(--ink4);
  letter-spacing: 1px;
  text-transform: uppercase;
  margin-top: 3px;
}

.ledger-col.cr .ledger-val { color: var(--green); }
.ledger-col.db .ledger-val { color: var(--red); }

/* ── SECTION HEADERS ── */
.section-head {
  display: flex;
  align-items: center;
  gap: 0;
  margin: 14px 16px 8px;
}

.sh-line { flex: 1; height: 1px; background: var(--rule); }
.sh-text {
  font-family: 'Playfair Display', serif;
  font-size: 11px;
  font-weight: 700;
  letter-spacing: 3px;
  text-transform: uppercase;
  color: var(--ink);
  padding: 0 10px;
}

/* ── CARDS / PANELS ── */
.panel {
  margin: 0 16px 14px;
  border: 1.5px solid var(--rule);
  background: var(--paper2);
}

.panel-head {
  border-bottom: 1px solid var(--rule);
  padding: 8px 14px;
  font-family: 'Playfair Display', serif;
  font-size: 13px;
  font-weight: 700;
  color: var(--ink);
  letter-spacing: 0.5px;
  background: var(--paper3);
  display: flex;
  align-items: center;
  justify-content: space-between;
}

.panel-body { padding: 14px; }

/* ── LABELS ── */
label {
  font-family: 'IM Fell English', serif;
  font-style: italic;
  font-size: 11px;
  color: var(--ink3);
  display: block;
  margin-bottom: 5px;
  letter-spacing: 0.5px;
}

/* ── TIPO TOGGLE ── */
.tipo { display: grid; grid-template-columns: 1fr 1fr; gap: 0; margin-bottom: 12px; border: 1.5px solid var(--rule); }

.tb {
  padding: 11px;
  border: none;
  border-right: 1px solid var(--rule);
  background: var(--paper2);
  font-family: 'Playfair Display', serif;
  font-size: 13px;
  font-weight: 700;
  cursor: pointer;
  text-align: center;
  color: var(--ink3);
  transition: all 0.15s ease;
  letter-spacing: 0.5px;
}

.tb:last-child { border-right: none; }

.tb:hover { background: var(--paper3); }

.tb.gasto.sel  {
  background: var(--ink);
  color: var(--paper);
}
.tb.ingreso.sel {
  background: var(--green);
  color: var(--paper);
}

/* ── INPUTS ── */
.row2 { display: grid; grid-template-columns: 1fr 90px; gap: 8px; margin-bottom: 10px; }

input {
  width: 100%;
  padding: 10px 12px;
  border: 1.5px solid var(--ink3);
  border-radius: 0;
  font-size: 15px;
  color: var(--ink);
  background: var(--paper);
  outline: none;
  -webkit-appearance: none;
  font-family: 'IM Fell English', Georgia, serif;
  transition: all 0.15s ease;
}

input:focus {
  border-color: var(--ink);
  background: #FFF8E8;
  box-shadow: inset 0 1px 3px rgba(0,0,0,0.1);
}

input::placeholder {
  color: var(--faded);
  font-style: italic;
}

/* ── QUICK EXAMPLES ── */
.exs { display: flex; flex-wrap: wrap; gap: 5px; margin-bottom: 12px; }

.ex {
  padding: 5px 10px;
  background: transparent;
  color: var(--ink3);
  border: 1px solid var(--ink4);
  font-family: 'IM Fell English', serif;
  font-size: 11px;
  font-style: italic;
  cursor: pointer;
  transition: all 0.15s ease;
}

.ex:hover {
  background: var(--ink);
  color: var(--paper);
  border-color: var(--ink);
  transform: translateY(-1px);
}

.ex:active { transform: scale(0.97); }

/* ── CATEGORY BUTTONS ── */
.cats { display: grid; grid-template-columns: repeat(4,1fr); gap: 5px; margin-bottom: 12px; }

.cb {
  padding: 8px 3px;
  border: 1px solid var(--ink4);
  background: var(--paper);
  font-family: 'IM Fell English', serif;
  font-size: 10px;
  font-style: italic;
  text-align: center;
  cursor: pointer;
  color: var(--ink3);
  transition: all 0.15s ease;
}

.cb:hover {
  background: var(--paper3);
  border-color: var(--ink);
  transform: translateY(-1px);
  box-shadow: 2px 2px 0 var(--ink4);
}

.cb.sel {
  background: var(--ink);
  color: var(--paper);
  border-color: var(--ink);
}

.cb span { display: block; font-size: 15px; margin-bottom: 2px; font-style: normal; }

/* ── SUBMIT BUTTON ── */
.btn {
  width: 100%;
  padding: 14px;
  background: var(--ink);
  color: var(--paper);
  border: 2px solid var(--ink);
  font-family: 'Playfair Display', serif;
  font-size: 14px;
  font-weight: 700;
  letter-spacing: 2px;
  text-transform: uppercase;
  cursor: pointer;
  transition: all 0.2s ease;
}

.btn:hover {
  background: var(--ink2);
  transform: translateY(-2px);
  box-shadow: 3px 3px 0 var(--ink4);
}

.btn:active { transform: translateY(0); box-shadow: none; }
.btn:disabled { background: var(--faded); border-color: var(--faded); cursor: not-allowed; transform: none; box-shadow: none; }

/* ── TOAST ── */
.toast {
  position: fixed;
  bottom: 24px;
  left: 50%;
  transform: translateX(-50%) translateY(120px);
  background: var(--ink);
  color: var(--paper);
  padding: 11px 22px;
  border: 1px solid var(--paper3);
  font-family: 'IM Fell English', serif;
  font-size: 13px;
  font-style: italic;
  transition: transform 0.3s ease;
  z-index: 100;
  white-space: nowrap;
  box-shadow: 4px 4px 0 var(--ink3);
}

.toast.show { transform: translateX(-50%) translateY(0); }
.toast.ok  { background: var(--green); }
.toast.err { background: var(--red); }

/* ── TRANSACTION ROWS ── */
.txn {
  display: flex;
  align-items: center;
  padding: 10px 14px;
  border-bottom: 1px solid var(--paper3);
  gap: 12px;
  transition: background 0.15s ease;
  cursor: default;
}

.txn:hover { background: var(--paper); }
.txn:last-child { border-bottom: none; }

.ti {
  width: 34px; height: 34px;
  border: 1px solid var(--ink4);
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 16px;
  flex-shrink: 0;
  background: var(--paper);
  transition: transform 0.15s ease;
}

.txn:hover .ti { transform: scale(1.1); }

.td { flex: 1; min-width: 0; }

.tn {
  font-family: 'Playfair Display', serif;
  font-size: 13px;
  font-weight: 700;
  color: var(--ink);
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}

.tm {
  font-family: 'IM Fell English', serif;
  font-style: italic;
  font-size: 10px;
  color: var(--ink4);
  margin-top: 1px;
}

.ta {
  font-family: 'Playfair Display', serif;
  font-size: 14px;
  font-weight: 700;
  flex-shrink: 0;
}

.ta.g { color: var(--green); }
.ta.r { color: var(--red); }

.empty {
  padding: 28px 16px;
  text-align: center;
  color: var(--faded);
  font-family: 'IM Fell English', serif;
  font-style: italic;
  font-size: 14px;
}

.empty big { display: block; font-size: 32px; margin-bottom: 10px; font-style: normal; }

/* ── RULE ORNAMENT ── */
.ornament {
  text-align: center;
  color: var(--ink4);
  font-size: 14px;
  margin: 4px 0 10px;
  letter-spacing: 6px;
}
</style>
</head>
<body>

<!-- MASTHEAD -->
<div class="masthead">
  <div class="mast-top">
    <div class="mast-logo">
      <svg class="logo-svg" viewBox="0 0 80 80" fill="none" xmlns="http://www.w3.org/2000/svg">
        <circle cx="40" cy="40" r="34" stroke="#0F0A04" stroke-width="1.5" fill="none"/>
        <ellipse cx="40" cy="40" rx="34" ry="13" stroke="#0F0A04" stroke-width="1" fill="none" opacity="0.5"/>
        <ellipse cx="40" cy="40" rx="34" ry="22" stroke="#0F0A04" stroke-width="1" fill="none" opacity="0.35"/>
        <ellipse cx="40" cy="40" rx="19" ry="34" stroke="#0F0A04" stroke-width="1" fill="none" opacity="0.45"/>
        <ellipse cx="40" cy="40" rx="7"  ry="34" stroke="#0F0A04" stroke-width="1" fill="none" opacity="0.35"/>
        <line x1="6" y1="40" x2="74" y2="40" stroke="#0F0A04" stroke-width="0.8" opacity="0.3"/>
        <line x1="40" y1="6"  x2="40" y2="74" stroke="#0F0A04" stroke-width="0.8" opacity="0.3"/>
        <text x="40" y="47" font-family="Georgia,serif" font-size="19" font-weight="700" fill="#0F0A04" text-anchor="middle" letter-spacing="1">MN</text>
      </svg>
      <div>
        <div class="mast-brand">MN</div>
        <div class="mast-sub">Personal Finance Ledger</div>
      </div>
    </div>
    <div class="mast-date" id="fecha-mast"></div>
  </div>
  <hr class="mast-rule">
  <div class="mast-headline">Est. MMXXVI &nbsp;·&nbsp; Registro de Cuentas &nbsp;·&nbsp; Edición Personal</div>
</div>

<!-- LEDGER STATS -->
<div class="ledger-bar">
  <div class="ledger-col cr">
    <div class="ledger-val" id="si">—</div>
    <div class="ledger-lbl">Haber</div>
  </div>
  <div class="ledger-divider"></div>
  <div class="ledger-col db">
    <div class="ledger-val" id="sg">—</div>
    <div class="ledger-lbl">Debe</div>
  </div>
  <div class="ledger-divider"></div>
  <div class="ledger-col">
    <div class="ledger-val" id="ss">—</div>
    <div class="ledger-lbl">Saldo</div>
  </div>
</div>

<div class="ornament">— ✦ —</div>

<!-- NUEVA TRANSACCIÓN -->
<div class="section-head">
  <div class="sh-line"></div>
  <div class="sh-text">Nuevo Asiento</div>
  <div class="sh-line"></div>
</div>

<div class="panel">
  <div class="panel-head">
    <span>Registrar Movimiento</span>
    <span style="font-style:italic;font-weight:400;font-size:11px;color:var(--ink4)">La IA clasifica automáticamente</span>
  </div>
  <div class="panel-body">

    <label>Naturaleza del movimiento</label>
    <div class="tipo">
      <button class="tb gasto sel" onclick="setTipo('Gasto')">↓ &nbsp;Cargo / Gasto</button>
      <button class="tb ingreso" onclick="setTipo('Ingreso')">↑ &nbsp;Abono / Ingreso</button>
    </div>

    <label>Concepto e importe</label>
    <div class="row2">
      <input type="text" id="desc" placeholder="pan, gasolina, salario…" autocomplete="off" autocorrect="off" spellcheck="false">
      <input type="number" id="monto" placeholder="€" step="0.01" min="0">
    </div>

    <div class="exs">
      <button class="ex" onclick="q('pan',1)">🍞 Pan</button>
      <button class="ex" onclick="q('mercadona',65)">🛒 Mercadona</button>
      <button class="ex" onclick="q('gasolina',48)">⛽ Gasolina</button>
      <button class="ex" onclick="q('gym',35)">🏋️ Gym</button>
      <button class="ex" onclick="q('netflix',12.99)">📺 Netflix</button>
      <button class="ex" onclick="q('salario',1500,true)">💼 Salario</button>
    </div>

    <label>Partida contable <span style="font-weight:400">(opcional)</span></label>
    <div class="cats">
      <button class="cb" data-cat="Comida"      onclick="setCat(this)"><span>🍽️</span>Comida</button>
      <button class="cb" data-cat="Hogar"       onclick="setCat(this)"><span>🏠</span>Hogar</button>
      <button class="cb" data-cat="Transporte"  onclick="setCat(this)"><span>🚗</span>Transp.</button>
      <button class="cb" data-cat="Salud"       onclick="setCat(this)"><span>💊</span>Salud</button>
      <button class="cb" data-cat="Ocio"        onclick="setCat(this)"><span>🎉</span>Ocio</button>
      <button class="cb" data-cat="Inversiones" onclick="setCat(this)"><span>📈</span>Invers.</button>
      <button class="cb" data-cat="Trabajo"     onclick="setCat(this)"><span>💼</span>Trabajo</button>
      <button class="cb" data-cat="Otros"       onclick="setCat(this)"><span>📦</span>Otros</button>
    </div>

    <button class="btn" id="btn" onclick="enviar()">Asentar en el Libro</button>
  </div>
</div>

<!-- HISTORIAL -->
<div class="section-head">
  <div class="sh-line"></div>
  <div class="sh-text">Últimos Asientos</div>
  <div class="sh-line"></div>
</div>

<div class="panel">
  <div class="panel-head">
    <span>Libro Mayor</span>
    <span id="cnt" style="font-style:italic;font-weight:400;font-size:11px;color:var(--ink4)"></span>
  </div>
  <div id="list"><div class="empty"><big>✦</big>Sin asientos registrados</div></div>
</div>

<div class="toast" id="toast"></div>

<script>
const ICONS={Comida:'🍽️',Hogar:'🏠',Transporte:'🚗',Salud:'💊',Ocio:'🎉',Inversiones:'📈',Trabajo:'💼',Ingresos:'💰',Otros:'📦'};
let tipo='Gasto', cat=null;

const now = new Date();
const dias = ['Domingo','Lunes','Martes','Miércoles','Jueves','Viernes','Sábado'];
const meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];
document.getElementById('fecha-mast').innerHTML =
  dias[now.getDay()] + ', ' + now.getDate() + ' de ' + meses[now.getMonth()] + ' de ' + now.getFullYear();

function setTipo(t) {
  tipo = t;
  document.querySelectorAll('.tb').forEach(b => b.classList.remove('sel'));
  document.querySelector('.tb.' + t.toLowerCase()).classList.add('sel');
}

function setCat(b) {
  document.querySelectorAll('.cb').forEach(x => x.classList.remove('sel'));
  if (cat === b.dataset.cat) { cat = null; }
  else { cat = b.dataset.cat; b.classList.add('sel'); }
}

function q(d, m, ing=false) {
  document.getElementById('desc').value = d;
  document.getElementById('monto').value = m;
  if (ing) setTipo('Ingreso'); else setTipo('Gasto');
}

function fmt(n) {
  if (n === 0) return '0,00 €';
  return n.toLocaleString('es-ES', {minimumFractionDigits:2, maximumFractionDigits:2}) + ' €';
}

function toast(msg, type) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className = 'toast ' + type; t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 3200);
}

async function stats() {
  try {
    const r = await fetch('/api/stats');
    const d = await r.json();
    document.getElementById('si').textContent = fmt(d.ingresos);
    document.getElementById('sg').textContent = fmt(d.gastos);
    const s = d.ingresos - d.gastos;
    document.getElementById('ss').textContent = (s < 0 ? '-' : '') + fmt(Math.abs(s));
    document.getElementById('ss').style.color = s >= 0 ? 'var(--green)' : 'var(--red)';
  } catch(e) {}
}

async function hist() {
  try {
    const r = await fetch('/api/transacciones?limit=15');
    const d = await r.json();
    const txns = d.transacciones || [];
    document.getElementById('cnt').textContent = txns.length + ' entradas';
    if (!txns.length) {
      document.getElementById('list').innerHTML = '<div class="empty"><big>✦</big>El libro mayor está vacío.<br>Registre el primer asiento.</div>';
      return;
    }
    document.getElementById('list').innerHTML = txns.map(t => `
      <div class="txn">
        <div class="ti">${ICONS[t.categoria] || '📦'}</div>
        <div class="td">
          <div class="tn">${t.descripcion}</div>
          <div class="tm">${t.categoria} &nbsp;·&nbsp; ${t.fecha}</div>
        </div>
        <div class="ta ${t.tipo === 'Ingreso' ? 'g' : 'r'}">${t.tipo === 'Ingreso' ? '+' : '-'}${parseFloat(t.monto).toLocaleString('es-ES',{minimumFractionDigits:2})} €</div>
      </div>`).join('');
  } catch(e) {
    document.getElementById('list').innerHTML = '<div class="empty"><big>⚠</big>Error de conexión</div>';
  }
}

async function enviar() {
  const desc  = document.getElementById('desc').value.trim();
  const monto = parseFloat(document.getElementById('monto').value);
  if (!desc)  { toast('Indique un concepto', 'err'); return; }
  if (!monto || monto <= 0) { toast('Indique un importe válido', 'err'); return; }
  const btn = document.getElementById('btn');
  btn.disabled = true; btn.textContent = 'Registrando…';
  try {
    const body = { descripcion: desc, monto, tipo };
    if (cat) body.categoria = cat;
    const r = await fetch('/api/transaccion', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(body)
    });
    const d = await r.json();
    if (d.ok) {
      toast(`Asentado — ${d.categoria} · ${monto.toLocaleString('es-ES',{minimumFractionDigits:2})} €`, 'ok');
      document.getElementById('desc').value = '';
      document.getElementById('monto').value = '';
      cat = null;
      document.querySelectorAll('.cb').forEach(b => b.classList.remove('sel'));
      await stats(); await hist();
    } else toast('Error: ' + (d.error || 'desconocido'), 'err');
  } catch(e) { toast('Sin conexión', 'err'); }
  finally { btn.disabled = false; btn.textContent = 'Asentar en el Libro'; }
}

['desc','monto'].forEach(id =>
  document.getElementById(id).addEventListener('keypress', e => { if (e.key === 'Enter') enviar(); })
);

stats(); hist();
</script>
</body>
</html>"""


@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/api/transaccion", methods=["POST"])
def api_nueva():
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "Sin datos"}), 400
    descripcion  = data.get("descripcion", "").strip()
    monto        = float(data.get("monto", 0))
    tipo_forzado = data.get("tipo")
    cat_forzada  = data.get("categoria")
    if not descripcion or monto <= 0:
        return jsonify({"ok": False, "error": "Datos inválidos"}), 400
    all_cats = list(REGLAS.keys()) + ["Otros"]
    if cat_forzada and cat_forzada in all_cats:
        categoria = cat_forzada
        tipo = tipo_forzado or ("Ingreso" if categoria == "Ingresos" else "Gasto")
    else:
        categoria, tipo = clasificar(descripcion)
        if tipo_forzado in ["Ingreso", "Gasto"]:
            tipo = tipo_forzado
            if tipo == "Ingreso":
                categoria = "Ingresos"
    fila = añadir_al_excel(descripcion, monto, categoria, tipo)
    return jsonify({"ok": True, "descripcion": descripcion, "monto": monto,
                    "categoria": categoria, "tipo": tipo, "fila": fila})

@app.route("/api/stats")
def api_stats():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Transacciones"]
        ingresos = gastos = 0.0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[1] and isinstance(row[1], (int, float)):
                if row[4] == "Ingreso":
                    ingresos += row[1]
                else:
                    gastos += row[1]
        return jsonify({"ingresos": round(ingresos,2), "gastos": round(gastos,2)})
    except Exception as e:
        return jsonify({"ingresos": 0, "gastos": 0, "error": str(e)})

@app.route("/api/transacciones")
def api_transacciones():
    limit = int(request.args.get("limit", 20))
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Transacciones"]
        txns = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            desc, monto, fecha, cat, tipo = row[0], row[1], row[2], row[3], row[4]
            if desc and monto:
                txns.append({"descripcion": str(desc), "monto": float(monto),
                             "fecha": str(fecha) if fecha else "",
                             "categoria": str(cat) if cat else "Otros",
                             "tipo": str(tipo) if tipo else "Gasto"})
        txns.reverse()
        return jsonify({"transacciones": txns[:limit]})
    except Exception as e:
        return jsonify({"transacciones": [], "error": str(e)})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
